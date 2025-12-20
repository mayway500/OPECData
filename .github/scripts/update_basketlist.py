#!/usr/bin/env python3
"""
Update 'Basketlist' sheet from 'oilpricechart' in the given XLSX (if present),
then export both sheets ('oilpricechart' and 'Basketlist') to CSV files.

Usage:
    python .github/scripts/update_and_export.py path/to/Opecpricechart.xlsx path/to/export_dir
"""
import sys
import os
import csv
from openpyxl import load_workbook

def clear_sheet(ws):
    # Unmerge all merged cells then clear values
    merges = list(ws.merged_cells.ranges)
    for m in merges:
        try:
            ws.unmerge_cells(str(m))
        except Exception:
            pass
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None

def copy_sheet_values(src, dst):
    # Clear destination then copy cell values and merged ranges and column widths (best effort)
    clear_sheet(dst)

    max_row = src.max_row or 0
    max_col = src.max_column or 0

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            dst.cell(row=r, column=c).value = src.cell(row=r, column=c).value

    # copy merged ranges
    for m in src.merged_cells.ranges:
        try:
            dst.merge_cells(str(m))
        except Exception:
            pass

    # copy column widths (best-effort)
    try:
        for col_letter, dim in src.column_dimensions.items():
            if dim.width:
                dst.column_dimensions[col_letter].width = dim.width
    except Exception:
        pass

def sheet_to_csv(ws, csv_path):
    # Write sheet values to CSV. Use UTF-8-sig so Excel on Windows reads UTF-8 nicely.
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    with open(csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, quoting=csv.QUOTE_MINIMAL)
        for r in range(1, max_row + 1):
            row_vals = []
            for c in range(1, max_col + 1):
                val = ws.cell(row=r, column=c).value
                if val is None:
                    row_vals.append('')
                else:
                    row_vals.append(val)
            writer.writerow(row_vals)

def main(xlsx_path, export_dir):
    if not os.path.isfile(xlsx_path):
        print(f"Workbook not found: {xlsx_path}")
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True)  # data_only to read values instead of formulas
    src_name = 'oilpricechart'
    dst_name = 'Basketlist'

    # Update Basketlist from oilpricechart if possible
    if src_name in wb.sheetnames:
        src = wb[src_name]
        if dst_name in wb.sheetnames:
            dst = wb[dst_name]
        else:
            dst = wb.create_sheet(dst_name)
        print(f"Copying values from '{src_name}' to '{dst_name}'...")
        # Need to reload workbook in write mode to preserve changes; openpyxl allows saving the same wb even if data_only=True
        # But to ensure we have a writeable workbook, re-open without data_only.
        wb_write = load_workbook(xlsx_path)
        src_write = wb_write[src_name]
        if dst_name in wb_write.sheetnames:
            dst_write = wb_write[dst_name]
        else:
            dst_write = wb_write.create_sheet(dst_name)
        copy_sheet_values(src_write, dst_write)
        wb_write.save(xlsx_path)
        # reload workbook for exporting (read the up-to-date values)
        wb = load_workbook(xlsx_path, data_only=True)
    else:
        print(f"Source sheet '{src_name}' not found. Skipping update of '{dst_name}'.")

    # Ensure export directory exists
    os.makedirs(export_dir, exist_ok=True)

    # Export the sheets to CSV
    to_export = []
    if src_name in wb.sheetnames:
        to_export.append(src_name)
    if dst_name in wb.sheetnames:
        to_export.append(dst_name)

    for sheet_name in to_export:
        ws = wb[sheet_name]
        # create a safe filename, lower-case for convention
        filename = f"{sheet_name}.csv"
        csv_path = os.path.join(export_dir, filename)
        print(f"Exporting sheet '{sheet_name}' to {csv_path} ...")
        sheet_to_csv(ws, csv_path)

    print("Export complete.")

if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: update_and_export.py path/to/Opecpricechart.xlsx path/to/export_dir")
        sys.exit(2)
    xlsx_path = sys.argv[1]
    export_dir = sys.argv[2]
    main(xlsx_path, export_dir)
